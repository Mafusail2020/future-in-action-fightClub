"""City budget of the Zhytomyr territorial community.

openbudget.gov.ua is geo-fenced (Akamai), so the numbers come from the city's
own open data on data.gov.ua: the budget department publishes council decisions
with annexes, where «Додаток 1» is incomes and «Додаток 3» is expenses as XLSX
inside a ZIP (file names inside are cp1251-encoded).

fetch_site returns a JSON-friendly dict of parsed rows (that is what gets
cached); parse() turns them into city-wide metrics + one readable document.
"""

from __future__ import annotations

import io
import re
import zipfile
from datetime import UTC, datetime
from typing import Any

from ingestion.common import http
from ingestion.common.api_scraper import ApiScraperSource
from ingestion.common.base_source import RawDoc, SourceOutput

PACKAGE_ID = "d85474f8-5459-44a3-a6d5-4f40efd98b76"  # «Бюджет Житомирської МТГ»
PACKAGE_URL = "https://data.gov.ua/api/3/action/package_show"
DATASET_PAGE = f"https://data.gov.ua/dataset/{PACKAGE_ID}"

# income classification code -> short name (top level of «Додаток 1»)
_INCOME_GROUPS = {
    "10000000": "податкові надходження",
    "20000000": "неподаткові надходження",
    "30000000": "доходи від операцій з капіталом",
    "40000000": "офіційні трансферти",
    "50000000": "цільові фонди",
}


def _decision_year(name: str) -> int | None:
    """Budget decisions are titled «... Про бюджет ... на YYYY рік»."""
    match = re.search(r"на\s+(20\d\d)\s+рік", name, re.IGNORECASE)
    return int(match.group(1)) if match else None


def _annex_number(filename: str) -> int | None:
    """Digits before the extension survive the cp1251 mojibake of inner names."""
    match = re.search(r"(\d+)\s*\.xlsx$", filename, re.IGNORECASE)
    return int(match.group(1)) if match else None


def _sheet_rows(data: bytes) -> list[list[Any]]:
    from openpyxl import load_workbook  # local import: heavy, scraper-only dep

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    rows = [list(row) for row in workbook.active.iter_rows(values_only=True)]
    workbook.close()
    return rows


def _parse_incomes(rows: list[list[Any]]) -> list[dict[str, Any]]:
    """Rows of «Додаток 1»: code | name | total | general fund | special fund."""
    out = []
    for row in rows:
        code = str(row[0]).strip() if row and row[0] is not None else ""
        if code in _INCOME_GROUPS or code.lower() in ("всього", "разом"):
            values = [c for c in row[2:] if isinstance(c, (int, float))]
            if values:
                out.append({"code": code, "name": _INCOME_GROUPS.get(code, "всього"),
                            "total": float(values[0])})
    return out


def _parse_expenses(rows: list[list[Any]]) -> list[dict[str, Any]]:
    """Rows of «Додаток 3»: top-level managers are codes like 0200000 with the
    grand total in the last numeric column. The name sits in the first string
    cell after the code — its exact column shifts with merged-cell layout."""
    out = []
    for row in rows:
        code = str(row[0]).strip() if row and row[0] is not None else ""
        if not re.fullmatch(r"\d{2}00000", code):
            continue
        name = next((c.strip() for c in row[1:5] if isinstance(c, str) and c.strip()), None)
        values = [c for c in row if isinstance(c, (int, float))]
        if name and values:
            out.append({"code": code, "name": name, "total": float(values[-1])})
    return out


class CityBudgetSource(ApiScraperSource):
    name = "Бюджет Житомирської МТГ (data.gov.ua)"
    cache_key = "citybudget_zhytomyr"
    ttl_hours = 24 * 30

    def fetch_site(self) -> Any:
        package = http.get_json(PACKAGE_URL, params={"id": PACKAGE_ID})["result"]
        resources = package.get("resources", [])

        # Newest budget decision, then the annexes ZIP with the same number.
        decisions = sorted(
            (r for r in resources if _decision_year(r.get("name", ""))),
            key=lambda r: _decision_year(r["name"]),
            reverse=True,
        )
        if not decisions:
            raise RuntimeError("no budget decision resource found in the dataset")
        decision = decisions[0]
        number = re.search(r"№\s*(\d+)", decision["name"])
        zips = [
            r for r in resources
            if r.get("format", "").upper() == "ZIP"
            and (not number or number.group(1) in r.get("name", ""))
        ]
        if not zips:
            raise RuntimeError(f"no annexes ZIP for decision {decision['name']!r}")

        archive = zipfile.ZipFile(io.BytesIO(http.get_bytes(zips[0]["url"])),
                                  metadata_encoding="cp1251")
        incomes: list[dict[str, Any]] = []
        expenses: list[dict[str, Any]] = []
        for entry in archive.namelist():
            if "~$" in entry or not entry.lower().endswith(".xlsx"):
                continue
            annex = _annex_number(entry)
            if annex == 1:
                incomes = _parse_incomes(_sheet_rows(archive.read(entry)))
            elif annex == 3:
                expenses = _parse_expenses(_sheet_rows(archive.read(entry)))

        return {
            "decision": decision["name"],
            "year": _decision_year(decision["name"]),
            "incomes": incomes,
            "expenses": expenses,
        }

    def parse(self, raw: Any) -> SourceOutput:
        year = raw.get("year")
        incomes, expenses = raw.get("incomes", []), raw.get("expenses", [])
        if not incomes and not expenses:
            print("  ! citybudget: nothing parsed from annexes")
            return SourceOutput()

        metrics: list[dict[str, Any]] = []
        lines: list[str] = [f"Рішення: {raw.get('decision', '')}", ""]

        income_total = next((r["total"] for r in incomes if r["code"].lower() in ("всього", "разом")),
                            sum(r["total"] for r in incomes))
        metrics.append({"raion_slug": None, "metric": f"budget_income_total_{year}",
                        "value": income_total, "unit": "грн"})
        lines.append(f"Доходи бюджету громади на {year} рік: {income_total:,.0f} грн, з них:")
        for row in incomes:
            if row["code"] in _INCOME_GROUPS:
                metrics.append({
                    "raion_slug": None,
                    "metric": f"budget_income_{row['code'][:1]}0_{year}",
                    "value": row["total"], "unit": "грн",
                    "meta": {"name": row["name"], "code": row["code"]},
                })
                lines.append(f"- {row['name']}: {row['total']:,.0f} грн")

        top_spend = sorted(expenses, key=lambda r: r["total"], reverse=True)[:8]
        if top_spend:
            expense_total = sum(r["total"] for r in expenses)
            metrics.append({"raion_slug": None, "metric": f"budget_expense_total_{year}",
                            "value": expense_total, "unit": "грн"})
            lines += ["", f"Найбільші розпорядники видатків на {year} рік:"]
            lines += [f"- {r['name']}: {r['total']:,.0f} грн" for r in top_spend]

        doc = RawDoc(
            title=f"Бюджет Житомирської міської територіальної громади на {year} рік",
            content="\n".join(lines).replace(",", " "),  # 1 234 567 reads better in uk
            doc_type="budget",
            category=None,  # multi-topic: roads, transport, utilities all live here
            url=DATASET_PAGE,
            published_at=datetime.now(UTC).date(),
            external_id=f"citybudget-{year}",
            meta={"decision": raw.get("decision"), "year": year},
        )
        return SourceOutput(docs=[doc], metrics=metrics)
